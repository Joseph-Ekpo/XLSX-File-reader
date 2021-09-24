import csv
import time
from datetime import date, datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Dictionaries for storing key:value pairs --> {email1: UserID} & {email2: UserID}
match_Email1_ID = {}
match_Email2_ID = {}
previously_matched = []
email_lists = ['3 G W 8-3-21 - Copy.xlsx',
               '3 G W 8-9-21 - Copy.xlsx',
               '3 G W 8-23-21 - Copy.xlsx', # Previous week was skipped by naming accident
               '3 G W 8-30-21 - Copy.xlsx',
               '3 G W 8-31-21 - Copy.xlsx',
               '3 G W 9-7-21 - Copy.xlsx']



def compile_fullList():
    print()


# Open excel workbook from working directory
work_book = load_workbook(email_lists[5])
work_sheet = work_book['Sheet1']
work_sheet_TargetList = work_book['Sheet2']



# Iterate over sheet 1(emails sent to 3G customers) and get a list of all email address that got the 3G email
list_items = []

for row in range(2, 4248):
    for column in range(1, 2):
        char = get_column_letter(column)
        list_items.append(work_sheet[char + str(row)].value)


# Iterate over sheet 2 and get a list fill both dictionaries with data from 7/26 Target List (no duplicates)
for row in range(2, 13316):
    for column in range(1, 2):
        cell_reference = f"{get_column_letter(column)}{row}"
        nextVal = f"{get_column_letter(column + 1)}{row}"
        nextVal2 = f"{get_column_letter(column + 2)}{row}"

        # print(work_sheet_TargetList[cell_reference].value, "-->", work_sheet_TargetList[nextVal].value)
        match_Email1_ID[work_sheet_TargetList[cell_reference].value] = work_sheet_TargetList[nextVal].value

        second_email = work_sheet_TargetList[nextVal2].value

        if not second_email:
            pass
        else:
            # print(work_sheet_TargetList[nextVal2].value, "--->", work_sheet_TargetList[nextVal].value)
            match_Email2_ID[second_email] = work_sheet_TargetList[nextVal].value


# Check if this account has already been contacted in a previous list

# Test
# for key in match_Email1_ID:
#     print(key, match_Email1_ID[key])
#
# print("********************************************************************************************* Email 2 List  ****")
#
# for key in match_Email2_ID:
#     print(key, match_Email2_ID[key])


# Counter to match the # of emails sent
x = 0
duplicates = 0

wb = Workbook()
ws = wb.active
ws.title = "Emails - UserID"
ws.append(['Email', 'UserID'])
for email in list_items:
    x += 1
    if (email in match_Email1_ID) and (match_Email1_ID[email] not in previously_matched):
        previously_matched.append(match_Email1_ID[email])
        ws.append([email, match_Email1_ID[email]])
        print(email, "\n", match_Email1_ID[email], "\n") #  x, email, "------> ",
    elif (email in match_Email2_ID) and (match_Email2_ID[email] not in previously_matched):
        previously_matched.append(match_Email2_ID[email])
        ws.append([email, match_Email2_ID[email]])
        print(email, "\n", match_Email2_ID[email], "\n") #  x, email, "------> ",
    else:
        duplicates += 1


def get_time():
    cur_time = datetime.now()
    today = date.today()
    now = cur_time.strftime("%H:%M")

    dateTime_stamp = f"{today} | {now}"

    created_on = f"Updated on: {dateTime_stamp}"
    return created_on


# Create TimeStamp at the end of new file/sheet
ws.append([""])
ws.append(["Total Emails", x])
ws.append(["Emails w/ same ID", duplicates])
ws.append(["Total Recipients", x - duplicates])

ws.append([get_time()])

wb.save(f"Updated List_{email_lists[5]}")

print(f"Total Emails: {x}\nDuplicates: {duplicates}\nTotal Recipients: {x - duplicates}")

def get_upgrades():
    print()


def get_ID_upgrade(user_ID, SIM_ID, phone_Number):
    print()
